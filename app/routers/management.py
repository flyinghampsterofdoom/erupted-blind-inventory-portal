from __future__ import annotations

import csv
import json
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from decimal import InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path
from urllib.parse import urlencode

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import FileResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.auth import Principal, Role, is_admin_role, require_capability, require_role
from app.config import settings
from app.db import get_db
from app.dependencies import get_client_ip
from app.models import Campaign, CountGroup, CountSession, SessionStatus, Store, StoreForcedCount
from app.security.csrf import verify_csrf
from app.sync_square_campaigns import sync_campaigns
from app.services.audit_service import log_audit
from app.services.admin_store_count_service import (
    delete_draft_count as delete_admin_store_draft_count,
    get_draft_count as get_admin_store_draft_count,
    get_or_create_draft_count as get_or_create_admin_store_draft_count,
    list_active_store_rows as list_admin_store_count_stores,
    list_count_lines as list_admin_store_count_lines,
    list_draft_counts as list_admin_store_count_drafts,
    list_pushed_counts as list_admin_store_count_pushed,
    save_draft_count as save_admin_store_count_draft,
    submit_count as submit_admin_store_count,
)
from app.services.access_control_service import (
    allowed_dashboard_category_ids_for_role,
    fallback_allowed_for_role,
    list_access_control_settings,
    list_role_dashboard_category_access,
    permission_defs,
    principal_has_permission,
    save_role_dashboard_category_access,
    save_principal_permission_overrides,
    save_role_permission_overrides,
)
from app.services.change_box_count_service import ROLL_SIZES_BY_CODE
from app.services.change_box_count_service import delete_change_box_count, get_count_detail, list_counts_for_audit
from app.services.change_form_service import (
    DENOMS,
    get_change_form_detail,
    get_inventory_state,
    list_change_forms,
    submit_inventory_audit,
)
from app.services.cash_reconciliation_service import (
    get_cash_reconciliation_batch_detail,
    get_actual_cash_rows,
    get_expected_cash_by_day,
    list_cash_reconciliation_batches,
    list_square_enabled_stores,
    save_actual_cash_rows,
)
from app.services.customer_request_service import (
    add_item as add_customer_request_item,
    list_items_for_management as list_customer_request_items_for_management,
    list_submissions as list_customer_request_submissions,
    set_item_count as set_customer_request_item_count,
)
from app.services.cogs_report_service import build_cogs_report
from app.services.count_square_sync_service import (
    list_count_square_sync_report_rows,
    push_session_recount_variance_to_square,
    push_session_variance_to_square,
)
from app.services.count_group_audit_service import run_count_group_coverage_audit
from app.services.daily_chore_service import (
    DAILY_CHORE_SECTION_ORDER,
    add_global_task,
    delete_draft_sheet_for_management,
    delete_global_task,
    get_sheet_detail_for_audit,
    list_global_task_rows,
    list_sheets_for_audit,
    reorder_global_task,
)
from app.services.dashboard_layout_service import (
    build_dashboard_sections,
    create_dashboard_category,
    list_dashboard_layout_settings,
    save_dashboard_card_assignments,
    save_dashboard_categories,
)
from app.services.exchange_return_form_service import get_form_detail as get_exchange_return_form_detail
from app.services.exchange_return_form_service import list_forms as list_exchange_return_forms
from app.services.master_safe_audit_service import get_inventory_state as get_master_safe_inventory_state
from app.services.master_safe_audit_service import submit_audit as submit_master_safe_audit
from app.services.non_sellable_stock_take_service import (
    add_item as add_non_sellable_item,
    deactivate_item as deactivate_non_sellable_item,
    get_stock_take_detail,
    list_items as list_non_sellable_items,
    list_stock_takes_for_audit,
    unlock_stock_take,
)
from app.services.ordering_emergency_service import (
    add_sku_to_draft,
    build_emergency_editor_detail,
    create_emergency_draft,
    list_emergency_draft_history,
    push_emergency_draft,
    resolve_lookup_sku,
    save_draft_quantities,
)
from app.services.opening_checklist_service import get_submission_detail, list_submissions
from app.services.purchase_order_admin_service import (
    add_purchase_order_line_by_sku,
    autofill_square_variation_ids,
    delete_draft_purchase_order,
    generate_purchase_orders,
    get_purchase_order_detail,
    list_purchase_order_pdf_template_assignments,
    list_vendor_par_level_rows,
    list_active_vendors,
    list_purchase_orders,
    list_vendor_sku_configs,
    parse_generation_form,
    prefill_vendor_store_par_levels_from_living,
    save_purchase_order_pdf_template_assignments,
    update_purchase_order_pdf_template,
    save_vendor_store_par_levels,
    import_vendor_sku_configs_csv,
    receive_purchase_order,
    refresh_purchase_order_lines_from_catalog,
    save_purchase_order_lines,
    submit_purchase_order,
    upsert_vendor_sku_config,
)
from app.services.session_service import (
    create_management_user,
    create_count_group,
    create_forced_count,
    deactivate_count_group,
    get_management_variance_lines,
    group_management_data,
    list_management_users,
    list_store_login_rows,
    list_stores_with_rotation,
    purge_count_sessions,
    renumber_count_group_positions,
    reset_management_user_password,
    reset_manager_password,
    set_management_user_active,
    set_store_next_group,
    unlock_session,
    update_count_group,
    upsert_store_login_credentials,
)
from app.services.stock_value_on_hand_service import build_stock_value_on_hand_report
from app.services.square_vendor_service import sync_vendors_from_square
from app.services.square_ordering_data_service import sync_vendor_sku_configs_from_square
from app.services.store_par_reset_service import get_store_par_reset_data, save_store_par_levels

router = APIRouter(prefix='/management', tags=['management'])
management_access = require_capability('management.access', Role.ADMIN, Role.MANAGER, Role.LEAD)
admin_access = require_capability('management.admin', Role.ADMIN, Role.MANAGER)
groups_access = require_capability('management.groups', Role.ADMIN, Role.MANAGER)
users_access = require_capability('management.users', Role.ADMIN)


def _empty_emergency_editor_detail() -> dict:
    return {
        'vendors': [],
        'draft': None,
        'stores': [],
        'lookup_options': [],
        'rows': [],
    }


def _created_sort_key(row: dict) -> float:
    created_at = row.get('created_at')
    if isinstance(created_at, datetime):
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=timezone.utc)
        return created_at.timestamp()
    return 0.0


def _friendly_emergency_error(exc: Exception, *, action: str) -> str:
    raw = str(exc)
    lowered = raw.lower()
    if (
        ('undefinedtable' in lowered or 'relation' in lowered or '42p01' in lowered)
        and ('emergency_on_hand_drafts' in lowered or 'emergency_on_hand_draft_lines' in lowered)
    ):
        return (
            'Emergency On-Hand editor tables are missing in this environment. '
            'Run the Render schema update for emergency drafts, then retry.'
        )
    return f'Failed to {action}. Please retry or contact support if it continues.'


def _friendly_cash_reconciliation_error(exc: Exception, *, action: str) -> str:
    raw = str(exc)
    lowered = raw.lower()
    if 'undefinedtable' in lowered or 'undefinedcolumn' in lowered or '42p01' in lowered or '42703' in lowered:
        if (
            'cash_reconciliation_verification_batches' in lowered
            or 'cash_reconciliation_verifications' in lowered
            or 'batch_id' in lowered
        ):
            return (
                'Cash reconciliation schema is missing recent migration updates '
                '(verification batches). Run latest schema updates, then retry.'
            )
    return f'Failed to {action}. Please retry or contact support if it continues.'


@router.get('/home')
def home(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    role_defaults = {
        key: fallback_allowed_for_role(role=principal.role.value, permission_key=key)
        for key in ['management.access', 'management.admin', 'management.groups', 'management.users', 'store.access']
    }
    allowed_permission_keys = {
        key
        for key, fallback in role_defaults.items()
        if principal_has_permission(
            db,
            principal=principal,
            permission_key=key,
            fallback_allowed=fallback,
        )
    }
    sections = build_dashboard_sections(
        db,
        is_admin=is_admin_role(principal.role),
        allowed_permission_keys=allowed_permission_keys,
        allowed_category_ids=allowed_dashboard_category_ids_for_role(db, role=principal.role.value),
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_home.html',
        {
            'request': request,
            'principal': principal,
            'sections': sections,
            'can_manage_layout': 'management.admin' in allowed_permission_keys,
        },
    )


@router.get('/dashboard-settings')
def dashboard_settings_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    error = str(request.query_params.get('error', '')).strip() or None
    saved = str(request.query_params.get('saved', '')).strip() == '1'
    try:
        detail = list_dashboard_layout_settings(db)
    except SQLAlchemyError as exc:
        db.rollback()
        detail = {'categories': [], 'cards': []}
        error = error or f'Unable to load dashboard settings: {exc}'
    else:
        db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_dashboard_settings.html',
        {
            'request': request,
            'detail': detail,
            'error': error,
            'saved': saved,
        },
    )


@router.post('/dashboard-settings/categories')
async def dashboard_settings_create_category(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    __: None = Depends(verify_csrf),
):
    form = await request.form()
    name = str(form.get('name', '')).strip()
    try:
        create_dashboard_category(db, name=name)
    except ValueError as exc:
        db.rollback()
        query = urlencode({'error': str(exc)})
        return RedirectResponse(f'/management/dashboard-settings?{query}', status_code=303)
    db.commit()
    return RedirectResponse('/management/dashboard-settings?saved=1', status_code=303)


@router.post('/dashboard-settings/categories/save')
async def dashboard_settings_save_categories(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    __: None = Depends(verify_csrf),
):
    form = await request.form()
    rows: list[dict] = []
    for key in form.keys():
        key_text = str(key)
        if not key_text.startswith('name__'):
            continue
        category_id_raw = key_text.split('name__', 1)[1]
        if not category_id_raw.isdigit():
            continue
        category_id = int(category_id_raw)
        raw_position = str(form.get(f'position__{category_id}', '0')).strip()
        position = int(raw_position) if raw_position.lstrip('-').isdigit() else 0
        rows.append(
            {
                'id': category_id,
                'name': str(form.get(f'name__{category_id}', '')).strip(),
                'position': position,
                'active': str(form.get(f'active__{category_id}', '')).strip().lower() == 'on',
            }
        )
    try:
        save_dashboard_categories(db, rows=rows)
    except ValueError as exc:
        db.rollback()
        query = urlencode({'error': str(exc)})
        return RedirectResponse(f'/management/dashboard-settings?{query}', status_code=303)
    db.commit()
    return RedirectResponse('/management/dashboard-settings?saved=1', status_code=303)


@router.post('/dashboard-settings/cards/save')
async def dashboard_settings_save_cards(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    assignments: dict[str, int | None] = {}
    positions: dict[str, int] = {}
    for key in form.keys():
        key_text = str(key)
        if not key_text.startswith('category__'):
            continue
        card_key = key_text.split('category__', 1)[1]
        raw_category = str(form.get(key_text, '')).strip()
        category_id = int(raw_category) if raw_category.isdigit() else None
        assignments[card_key] = category_id
        raw_position = str(form.get(f'position__{card_key}', '')).strip()
        positions[card_key] = int(raw_position) if raw_position.lstrip('-').isdigit() else 9999
    save_dashboard_card_assignments(
        db,
        assignments=assignments,
        positions=positions,
        principal_id=principal.id,
    )
    db.commit()
    return RedirectResponse('/management/dashboard-settings?saved=1', status_code=303)


def _render_placeholder(request: Request, title: str) -> object:
    return request.app.state.templates.TemplateResponse(
        'management_placeholder.html',
        {
            'request': request,
            'title': title,
        },
    )


def _parse_reconciliation_dates(start_raw: str, end_raw: str) -> tuple[date, date]:
    try:
        start_date = date.fromisoformat(start_raw)
        end_date = date.fromisoformat(end_raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='start_date and end_date must be YYYY-MM-DD') from exc
    if end_date < start_date:
        raise HTTPException(status_code=400, detail='end_date must be on or after start_date')
    return start_date, end_date


def _parse_admin_store_count_quantities(form) -> dict[str, Decimal | None]:
    values: dict[str, Decimal | None] = {}
    for key in form.keys():
        if not str(key).startswith('counted__'):
            continue
        variation_id = str(key).split('counted__', 1)[1]
        raw_value = str(form.get(key, '')).strip()
        if not raw_value:
            values[variation_id] = None
            continue
        try:
            qty = Decimal(raw_value)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f'Invalid count value for {variation_id}') from exc
        if qty < 0:
            raise ValueError(f'Count cannot be negative for {variation_id}')
        values[variation_id] = qty
    return values


def _safe_excel_filename_part(value: str) -> str:
    trimmed = value.strip().replace(' ', '_')
    safe = ''.join(ch for ch in trimmed if ch.isalnum() or ch in {'-', '_'})
    return safe or 'store'


@router.get('/store-count')
def management_store_count_page(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    stores = list_admin_store_count_stores(db)
    draft_counts = list_admin_store_count_drafts(db)
    pushed_counts = list_admin_store_count_pushed(db)
    selected_store_id_raw = str(request.query_params.get('store_id', '')).strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    selected_count_id_raw = str(request.query_params.get('count_id', '')).strip()
    selected_count_id = int(selected_count_id_raw) if selected_count_id_raw.isdigit() else None

    count = None
    lines: list[dict] = []
    is_new_draft = False
    error = None
    if selected_count_id is not None:
        try:
            count = get_admin_store_draft_count(db, count_id=selected_count_id)
            selected_store_id = count.store_id
            lines = list_admin_store_count_lines(db, count_id=count.id)
        except (ValueError, RuntimeError) as exc:
            error = str(exc)
    elif selected_store_id is not None:
        try:
            count, is_new_draft = get_or_create_admin_store_draft_count(
                db,
                store_id=selected_store_id,
                principal_id=principal.id,
            )
            lines = list_admin_store_count_lines(db, count_id=count.id)
        except (ValueError, RuntimeError) as exc:
            error = str(exc)

    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_store_count.html',
        {
            'request': request,
            'principal': principal,
            'stores': stores,
            'draft_counts': draft_counts,
            'pushed_counts': pushed_counts,
            'selected_store_id': selected_store_id,
            'count': count,
            'lines': lines,
            'is_new_draft': is_new_draft,
            'error': error,
            'save_ok': str(request.query_params.get('save_ok', '')) == '1',
            'submit_ok': str(request.query_params.get('submit_ok', '')) == '1',
            'submit_error': str(request.query_params.get('submit_error', '')).strip(),
        },
    )


@router.post('/store-count/{count_id}/save')
async def management_store_count_save(
    count_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    employee_name = str(form.get('employee_name', '')).strip()
    store_id_raw = str(form.get('store_id', '')).strip()
    redirect_store_id = int(store_id_raw) if store_id_raw.isdigit() else None
    try:
        counted_values = _parse_admin_store_count_quantities(form)
        count = get_admin_store_draft_count(db, count_id=count_id)
        save_admin_store_count_draft(
            db,
            count=count,
            employee_name=employee_name,
            counted_by_variation_id=counted_values,
            principal_id=principal.id,
        )
    except ValueError as exc:
        query = urlencode(
            {
                'store_id': redirect_store_id or '',
                'submit_error': str(exc),
            }
        )
        db.rollback()
        return RedirectResponse(f'/management/store-count?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ADMIN_STORE_COUNT_DRAFT_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'admin_store_count_id': count.id, 'store_id': count.store_id},
    )
    db.commit()
    query = urlencode({'store_id': count.store_id, 'save_ok': '1'})
    return RedirectResponse(f'/management/store-count?{query}', status_code=303)


@router.post('/store-count/{count_id}/submit')
async def management_store_count_submit(
    count_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    employee_name = str(form.get('employee_name', '')).strip()
    store_id_raw = str(form.get('store_id', '')).strip()
    redirect_store_id = int(store_id_raw) if store_id_raw.isdigit() else None
    try:
        counted_values = _parse_admin_store_count_quantities(form)
        count = get_admin_store_draft_count(db, count_id=count_id)
        result = submit_admin_store_count(
            db,
            count=count,
            employee_name=employee_name,
            counted_by_variation_id=counted_values,
            principal_id=principal.id,
        )
    except (ValueError, RuntimeError) as exc:
        query = urlencode(
            {
                'store_id': redirect_store_id or '',
                'submit_error': str(exc),
            }
        )
        db.commit()
        return RedirectResponse(f'/management/store-count?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ADMIN_STORE_COUNT_SUBMITTED_AND_PUSHED_TO_SQUARE',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'admin_store_count_id': count.id,
            'store_id': count.store_id,
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
        },
    )
    db.commit()
    query = urlencode({'store_id': count.store_id, 'submit_ok': '1'})
    return RedirectResponse(f'/management/store-count?{query}', status_code=303)


@router.post('/store-count/{count_id}/delete')
async def management_store_count_delete(
    count_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        store_id = delete_admin_store_draft_count(db, count_id=count_id)
    except ValueError as exc:
        query = urlencode({'submit_error': str(exc)})
        db.rollback()
        return RedirectResponse(f'/management/store-count?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ADMIN_STORE_COUNT_DRAFT_DELETED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'admin_store_count_id': count_id, 'store_id': store_id},
    )
    db.commit()
    return RedirectResponse('/management/store-count', status_code=303)


@router.post('/store-count/{count_id}/excel')
async def management_store_count_excel_download(
    count_id: int,
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    __: None = Depends(verify_csrf),
):
    form = await request.form()
    store_id_raw = str(form.get('store_id', '')).strip()
    redirect_store_id = int(store_id_raw) if store_id_raw.isdigit() else None
    employee_name = str(form.get('employee_name', '')).strip()
    try:
        counted_values = _parse_admin_store_count_quantities(form)
        count = get_admin_store_draft_count(db, count_id=count_id)
        lines = list_admin_store_count_lines(db, count_id=count.id)
    except ValueError as exc:
        query = urlencode(
            {
                'store_id': redirect_store_id or '',
                'submit_error': str(exc),
            }
        )
        return RedirectResponse(f'/management/store-count?{query}', status_code=303)

    store = db.execute(select(Store).where(Store.id == count.store_id)).scalar_one_or_none()
    store_name = str(store.name) if store else f'Store {count.store_id}'
    generated_at = datetime.now().astimezone()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Store Count Snapshot'
    ws.append(['Store Count (Full Inventory) Snapshot'])
    ws.append(['Store', store_name])
    ws.append(['Draft ID', count.id])
    ws.append(['Counter Name', employee_name])
    ws.append(['Generated At', generated_at.strftime('%Y-%m-%d %H:%M:%S %Z')])
    ws.append([])
    ws.append(['SKU', 'Item', 'Variation', 'Expected', 'Counted', 'Difference', 'Variation ID'])

    for line in lines:
        variation_id = str(line['variation_id'])
        expected = Decimal(str(line['expected_on_hand']))
        counted = counted_values.get(variation_id)
        variance = (counted - expected) if counted is not None else None
        ws.append(
            [
                line['sku'] or '',
                line['item_name'],
                line['variation_name'],
                float(expected),
                float(counted) if counted is not None else None,
                float(variance) if variance is not None else None,
                variation_id,
            ]
        )

    ws.freeze_panes = 'A8'
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 36

    for cell in ws['D'][7:]:
        cell.number_format = '0.000'
    for cell in ws['E'][7:]:
        cell.number_format = '0.000'
    for cell in ws['F'][7:]:
        cell.number_format = '0.000'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = generated_at.strftime('%Y%m%d_%H%M%S')
    store_part = _safe_excel_filename_part(store_name)
    filename = f'store_count_snapshot_{store_part}_{timestamp}.xlsx'
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@router.get('/cash-reconciliation')
def cash_reconciliation_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    today = date.today()
    default_start = (today - timedelta(days=6)).isoformat()
    default_end = today.isoformat()
    stores = list_square_enabled_stores(db)
    return request.app.state.templates.TemplateResponse(
        'management_cash_reconciliation.html',
        {
            'request': request,
            'stores': stores,
            'default_start_date': default_start,
            'default_end_date': default_end,
        },
    )


@router.get('/cash-reconciliation/expected')
def cash_reconciliation_expected(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    store_id_raw = str(request.query_params.get('store_id', '')).strip()
    start_raw = str(request.query_params.get('start_date', '')).strip()
    end_raw = str(request.query_params.get('end_date', '')).strip()
    if not store_id_raw.isdigit():
        raise HTTPException(status_code=400, detail='Valid store_id is required')
    if not start_raw or not end_raw:
        raise HTTPException(status_code=400, detail='start_date and end_date are required')
    start_date, end_date = _parse_reconciliation_dates(start_raw, end_raw)
    try:
        return get_expected_cash_by_day(
            db,
            store_id=int(store_id_raw),
            start_date=start_date,
            end_date=end_date,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=_friendly_cash_reconciliation_error(exc, action='load expected cash'),
        ) from exc


@router.get('/cash-reconciliation/actual')
def cash_reconciliation_actual(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    store_id_raw = str(request.query_params.get('store_id', '')).strip()
    start_raw = str(request.query_params.get('start_date', '')).strip()
    end_raw = str(request.query_params.get('end_date', '')).strip()
    if not store_id_raw.isdigit():
        raise HTTPException(status_code=400, detail='Valid store_id is required')
    if not start_raw or not end_raw:
        raise HTTPException(status_code=400, detail='start_date and end_date are required')
    start_date, end_date = _parse_reconciliation_dates(start_raw, end_raw)
    try:
        return get_actual_cash_rows(
            db,
            store_id=int(store_id_raw),
            start_date=start_date,
            end_date=end_date,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=_friendly_cash_reconciliation_error(exc, action='load actual cash'),
        ) from exc


@router.get('/cash-reconciliation/batches')
def cash_reconciliation_batches(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    store_id_raw = str(request.query_params.get('store_id', '')).strip()
    store_id: int | None = int(store_id_raw) if store_id_raw.isdigit() else None
    try:
        return list_cash_reconciliation_batches(
            db,
            store_id=store_id,
            limit=100,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=_friendly_cash_reconciliation_error(exc, action='load reconciliation batches'),
        ) from exc


@router.post('/cash-reconciliation/actual')
async def cash_reconciliation_actual_save(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    store_id_raw = str(form.get('store_id', '')).strip()
    if not store_id_raw.isdigit():
        raise HTTPException(status_code=400, detail='Valid store_id is required')
    store_id = int(store_id_raw)

    rows_json_raw = str(form.get('rows_json', '')).strip()
    if not rows_json_raw:
        raise HTTPException(status_code=400, detail='rows_json is required')
    try:
        parsed_rows = json.loads(rows_json_raw)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail='rows_json must be valid JSON') from exc
    if not isinstance(parsed_rows, list):
        raise HTTPException(status_code=400, detail='rows_json must be a list')

    expected_lookup: dict[date, int] = {}
    expected_json_raw = str(form.get('expected_json', '')).strip()
    if expected_json_raw:
        try:
            expected_rows = json.loads(expected_json_raw)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail='expected_json must be valid JSON') from exc
        if isinstance(expected_rows, list):
            for row in expected_rows:
                if not isinstance(row, dict):
                    continue
                raw_date = str(row.get('business_date') or '').strip()
                if not raw_date:
                    continue
                try:
                    business_date = date.fromisoformat(raw_date)
                except ValueError:
                    continue
                try:
                    expected_lookup[business_date] = int(row.get('expected_cash_cents'))
                except (TypeError, ValueError):
                    continue

    note = str(form.get('note', '')).strip() or None
    try:
        result = save_actual_cash_rows(
            db,
            store_id=store_id,
            principal_id=principal.id,
            rows=parsed_rows,
            expected_cash_by_date=expected_lookup,
            note=note,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=_friendly_cash_reconciliation_error(exc, action='save cash reconciliation'),
        ) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CASH_RECONCILIATION_ACTUAL_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'store_id': store_id, **result},
    )
    db.commit()
    return {'ok': True, **result}


@router.get('/cash-reconciliation/verification-batches/{batch_id}')
def cash_reconciliation_verification_batch_page(
    batch_id: int,
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_cash_reconciliation_batch_detail(
            db,
            batch_id=batch_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return request.app.state.templates.TemplateResponse(
        'management_cash_reconciliation_batch.html',
        {
            'request': request,
            'detail': detail,
        },
    )


@router.get('/store-par-reset')
def store_par_reset_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        data = get_store_par_reset_data(db, store_id=selected_store_id)
        load_error = None
    except Exception as exc:
        load_error = str(exc)
        data = {'stores': [], 'selected_store_id': None, 'change_box_rows': [], 'non_sellable_rows': []}
    return request.app.state.templates.TemplateResponse(
        'management_store_par_reset.html',
        {
            'request': request,
            'data': data,
            'saved': request.query_params.get('saved') == '1',
            'error_detail': load_error,
        },
    )


@router.post('/store-par-reset/save')
async def store_par_reset_save(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    store_id_raw = str(form.get('store_id', '')).strip()
    if not store_id_raw.isdigit():
        raise HTTPException(status_code=400, detail='Store is required')
    store_id = int(store_id_raw)

    change_box_par_by_code: dict[str, int] = {}
    change_box_level_by_code: dict[str, int] = {}
    non_sellable_par_by_item_id: dict[int, Decimal] = {}
    non_sellable_level_by_item_id: dict[int, Decimal] = {}

    for key, value in form.items():
        if key.startswith('cb_level__'):
            code = key.split('__', 1)[1]
            raw = str(value).strip()
            try:
                qty = int(raw) if raw else 0
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f'Invalid change box level for {code}') from exc
            change_box_level_by_code[code] = qty
            continue
        if key.startswith('cb_par__'):
            code = key.split('__', 1)[1]
            raw = str(value).strip()
            try:
                qty = int(raw) if raw else 0
            except ValueError as exc:
                raise HTTPException(status_code=400, detail=f'Invalid change box par for {code}') from exc
            change_box_par_by_code[code] = qty
            continue
        if key.startswith('ns_level__'):
            item_id_raw = key.split('__', 1)[1]
            if not item_id_raw.isdigit():
                continue
            raw = str(value).strip()
            try:
                qty = Decimal(raw.replace(',', '.')) if raw else Decimal('0.000')
            except (InvalidOperation, ValueError) as exc:
                raise HTTPException(status_code=400, detail='Invalid non-sellable level quantity') from exc
            non_sellable_level_by_item_id[int(item_id_raw)] = qty
            continue
        if key.startswith('ns_par__'):
            item_id_raw = key.split('__', 1)[1]
            if not item_id_raw.isdigit():
                continue
            raw = str(value).strip()
            try:
                qty = Decimal(raw.replace(',', '.')) if raw else Decimal('0.000')
            except (InvalidOperation, ValueError) as exc:
                raise HTTPException(status_code=400, detail='Invalid non-sellable par quantity') from exc
            non_sellable_par_by_item_id[int(item_id_raw)] = qty

    try:
        saved = save_store_par_levels(
            db,
            store_id=store_id,
            principal_id=principal.id,
            change_box_par_by_code=change_box_par_by_code,
            change_box_level_by_code=change_box_level_by_code,
            non_sellable_par_by_item_id=non_sellable_par_by_item_id,
            non_sellable_level_by_item_id=non_sellable_level_by_item_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='STORE_PAR_LEVELS_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata=saved,
    )
    db.commit()
    return RedirectResponse(f'/management/store-par-reset?store_id={store_id}&saved=1', status_code=303)


@router.get('/ordering-tool')
def ordering_tool_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    vendors = list_active_vendors(db)
    standard_orders = list_purchase_orders(db, limit=100)
    emergency_orders = list_emergency_draft_history(db, limit=100)
    orders: list[dict] = []
    for row in standard_orders:
        order_id = int(row['id'])
        status_text = str(getattr(row['status'], 'value', row['status']))
        orders.append(
            {
                'id': order_id,
                'display_id': str(order_id),
                'vendor_id': row.get('vendor_id'),
                'vendor_name': row.get('vendor_name'),
                'status': status_text,
                'created_at': row.get('created_at'),
                'submitted_at': row.get('submitted_at'),
                'open_href': f'/management/ordering-tool/orders/{order_id}',
                'receive_href': f'/management/ordering-tool/orders/{order_id}/receive',
                'discard_href': f'/management/ordering-tool/orders/{order_id}/delete',
                'can_receive': status_text == 'IN_TRANSIT',
                'can_discard': status_text in {'DRAFT', 'IN_TRANSIT'},
                'is_emergency': False,
            }
        )
    orders.extend(emergency_orders)
    orders.sort(key=_created_sort_key, reverse=True)
    orders = orders[:100]
    return request.app.state.templates.TemplateResponse(
        'management_ordering_tool.html',
        {
            'request': request,
            'vendors': vendors,
            'orders': orders,
            'default_reorder_weeks': settings.ordering_reorder_weeks_default,
            'default_stock_up_weeks': settings.ordering_stock_up_weeks_default,
            'default_history_lookback_days': settings.ordering_history_lookback_days_default,
            'query': request.query_params,
        },
    )


@router.get('/ordering-tool/emergency-editor')
def ordering_tool_emergency_editor_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    draft_raw = str(request.query_params.get('draft_id', '')).strip()
    draft_id = int(draft_raw) if draft_raw.isdigit() else None
    page_error = None
    try:
        detail = build_emergency_editor_detail(
            db,
            draft_id=draft_id,
        )
    except Exception as exc:
        db.rollback()
        detail = _empty_emergency_editor_detail()
        page_error = _friendly_emergency_error(exc, action='load emergency draft')
    return request.app.state.templates.TemplateResponse(
        'management_ordering_emergency_editor.html',
        {
            'request': request,
            'detail': detail,
            'query': request.query_params,
            'result': None,
            'page_error': page_error,
        },
    )


@router.post('/ordering-tool/emergency-editor/start-draft')
async def ordering_tool_emergency_editor_start_draft(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    vendor_raw = str(form.get('vendor_id', '')).strip()
    if not vendor_raw.isdigit():
        query = urlencode({'error': 'Select a vendor first'})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    try:
        draft = create_emergency_draft(
            db,
            vendor_id=int(vendor_raw),
            principal_id=principal.id,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    except Exception as exc:
        db.rollback()
        query = urlencode({'error': _friendly_emergency_error(exc, action='start emergency draft')})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    db.commit()
    query = urlencode({'draft_id': int(draft.id)})
    return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)


@router.post('/ordering-tool/emergency-editor/{draft_id}/add-sku')
async def ordering_tool_emergency_editor_add_sku(
    draft_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        draft = build_emergency_editor_detail(db, draft_id=draft_id).get('draft')
    except Exception as exc:
        db.rollback()
        query = urlencode({'error': _friendly_emergency_error(exc, action='open emergency draft')})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    if draft is None:
        query = urlencode({'error': 'Emergency draft not found'})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    vendor_id = int(draft.vendor_id)
    lookup = str(form.get('lookup', '')).strip()
    try:
        matched_sku = resolve_lookup_sku(
            db,
            vendor_id=vendor_id,
            lookup=lookup,
        )
        add_sku_to_draft(
            db,
            draft_id=draft_id,
            sku=matched_sku,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    except Exception as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': _friendly_emergency_error(exc, action='add SKU to emergency draft')})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    db.commit()
    query = urlencode({'draft_id': draft_id, 'added': matched_sku})
    return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)


def _parse_emergency_draft_quantities(form) -> dict[tuple[int, int], Decimal]:
    quantities: dict[tuple[int, int], Decimal] = {}
    for key, value in form.items():
        if not str(key).startswith('qty__'):
            continue
        try:
            _, line_raw, store_raw = str(key).split('__', 2)
            line_id = int(line_raw)
            store_id = int(store_raw)
        except Exception:
            continue
        raw = str(value).strip()
        if not raw:
            continue
        try:
            qty = Decimal(raw)
        except Exception:
            continue
        quantities[(line_id, store_id)] = qty
    return quantities


@router.post('/ordering-tool/emergency-editor/{draft_id}/save')
async def ordering_tool_emergency_editor_save(
    draft_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    quantities = _parse_emergency_draft_quantities(form)
    try:
        save_draft_quantities(
            db,
            draft_id=draft_id,
            quantities_by_line_store=quantities,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    except Exception as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': _friendly_emergency_error(exc, action='save emergency draft')})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_EMERGENCY_DRAFT_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'draft_id': draft_id, 'updated_cells': len(quantities)},
    )
    db.commit()
    query = urlencode({'draft_id': draft_id, 'saved': 1})
    return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)


@router.post('/ordering-tool/emergency-editor/{draft_id}/push')
async def ordering_tool_emergency_editor_push(
    draft_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    quantities = _parse_emergency_draft_quantities(form)
    try:
        save_draft_quantities(
            db,
            draft_id=draft_id,
            quantities_by_line_store=quantities,
        )
        result = push_emergency_draft(
            db,
            draft_id=draft_id,
            principal_id=principal.id,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)
    except Exception as exc:
        db.rollback()
        query = urlencode({'draft_id': draft_id, 'error': _friendly_emergency_error(exc, action='push emergency draft')})
        return RedirectResponse(f'/management/ordering-tool/emergency-editor?{query}', status_code=303)

    try:
        detail = build_emergency_editor_detail(
            db,
            draft_id=draft_id,
        )
    except Exception:
        db.rollback()
        detail = _empty_emergency_editor_detail()
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_EMERGENCY_ON_HAND_PUSHED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'draft_id': draft_id,
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
            'pushed': result['pushed'],
        },
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_ordering_emergency_editor.html',
        {
            'request': request,
            'detail': detail,
            'query': request.query_params,
            'result': result,
            'page_error': None,
        },
    )


@router.get('/ordering-tool/mappings')
def ordering_tool_mappings_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_vendor_raw = request.query_params.get('vendor_id', '').strip()
    selected_vendor_id = int(selected_vendor_raw) if selected_vendor_raw.isdigit() else None
    vendors = list_active_vendors(db)
    rows = list_vendor_sku_configs(db, vendor_id=selected_vendor_id)
    active_rows = [row for row in rows if row['active']]
    inactive_rows = [row for row in rows if not row['active']]
    return request.app.state.templates.TemplateResponse(
        'management_ordering_mappings.html',
        {
            'request': request,
            'vendors': vendors,
            'active_rows': active_rows,
            'inactive_rows': inactive_rows,
            'selected_vendor_id': selected_vendor_id,
            'query': request.query_params,
        },
    )


@router.get('/ordering-tool/par-levels')
def ordering_tool_par_levels_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    vendors = list_active_vendors(db)
    return request.app.state.templates.TemplateResponse(
        'management_ordering_par_levels.html',
        {
            'request': request,
            'vendors': vendors,
            'query': request.query_params,
        },
    )


@router.get('/ordering-tool/pdf-templates')
def ordering_tool_pdf_templates_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    detail = list_purchase_order_pdf_template_assignments(db)
    return request.app.state.templates.TemplateResponse(
        'management_ordering_pdf_templates.html',
        {
            'request': request,
            'detail': detail,
            'query': request.query_params,
        },
    )


@router.post('/ordering-tool/pdf-templates/save')
async def ordering_tool_pdf_templates_save(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    vendor_ids = [int(value) for value in form.getlist('vendor_ids') if str(value).strip().isdigit()]
    apply_generic = str(form.get('apply_generic', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
    try:
        touched = save_purchase_order_pdf_template_assignments(
            db,
            name=str(form.get('template_name', '')).strip(),
            legal_disclaimer=str(form.get('legal_disclaimer', '')).strip(),
            apply_generic=apply_generic,
            vendor_ids=vendor_ids,
            updated_by_principal_id=principal.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PDF_TEMPLATES_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'rows_touched': touched,
            'vendors_count': len(vendor_ids),
            'apply_generic': apply_generic,
        },
    )
    db.commit()
    query = urlencode({'saved': touched})
    return RedirectResponse(f'/management/ordering-tool/pdf-templates?{query}', status_code=303)


@router.post('/ordering-tool/pdf-templates/{template_id}/edit')
async def ordering_tool_pdf_templates_edit(
    template_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        updated = update_purchase_order_pdf_template(
            db,
            template_id=template_id,
            name=str(form.get('template_name', '')).strip(),
            legal_disclaimer=str(form.get('legal_disclaimer', '')).strip(),
            updated_by_principal_id=principal.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PDF_TEMPLATE_EDITED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'template_id': template_id, 'template_name': updated.name},
    )
    db.commit()
    query = urlencode({'edited': template_id})
    return RedirectResponse(f'/management/ordering-tool/pdf-templates?{query}', status_code=303)


@router.get('/ordering-tool/par-levels/{vendor_id}')
def ordering_tool_par_levels_vendor_page(
    vendor_id: int,
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    lookback_raw = request.query_params.get('history_lookback_days', '').strip()
    history_lookback_days = (
        int(lookback_raw)
        if lookback_raw.isdigit()
        else settings.ordering_history_lookback_days_default
    )
    try:
        detail = list_vendor_par_level_rows(
            db,
            vendor_id=vendor_id,
            history_lookback_days=history_lookback_days,
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return request.app.state.templates.TemplateResponse(
        'management_ordering_par_levels_vendor.html',
        {
            'request': request,
            'detail': detail,
            'history_lookback_days': history_lookback_days,
            'query': request.query_params,
        },
    )


@router.post('/ordering-tool/par-levels/{vendor_id}/save')
async def ordering_tool_par_levels_vendor_save(
    vendor_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    rows_raw = form.getlist('row_key')
    entries: list[tuple[int, str, int | None, int | None]] = []
    for row_key in rows_raw:
        parts = str(row_key).split('|', 1)
        if len(parts) != 2 or not parts[0].isdigit():
            continue
        store_id = int(parts[0])
        sku = parts[1].strip()
        manual_level_raw = str(form.get(f'manual_level__{row_key}', '')).strip()
        manual_par_raw = str(form.get(f'manual_par__{row_key}', '')).strip()
        manual_level = int(manual_level_raw) if manual_level_raw else None
        manual_par = int(manual_par_raw) if manual_par_raw else None
        entries.append((store_id, sku, manual_level, manual_par))
    try:
        saved = save_vendor_store_par_levels(
            db,
            vendor_id=vendor_id,
            entries=entries,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PAR_LEVELS_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'vendor_id': vendor_id, 'rows_saved': saved},
    )
    db.commit()
    query = urlencode({'saved': saved, 'history_lookback_days': form.get('history_lookback_days', '')})
    return RedirectResponse(f'/management/ordering-tool/par-levels/{vendor_id}?{query}', status_code=303)


@router.post('/ordering-tool/par-levels/{vendor_id}/prefill')
async def ordering_tool_par_levels_vendor_prefill(
    vendor_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    lookback_raw = str(form.get('history_lookback_days', '')).strip()
    history_lookback_days = (
        int(lookback_raw)
        if lookback_raw.isdigit()
        else settings.ordering_history_lookback_days_default
    )
    try:
        prefilled = prefill_vendor_store_par_levels_from_living(
            db,
            vendor_id=vendor_id,
            history_lookback_days=history_lookback_days,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PAR_LEVELS_PREFILLED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'vendor_id': vendor_id, 'rows_prefilled': prefilled, 'history_lookback_days': history_lookback_days},
    )
    db.commit()
    query = urlencode({'prefilled': prefilled, 'history_lookback_days': history_lookback_days})
    return RedirectResponse(f'/management/ordering-tool/par-levels/{vendor_id}?{query}', status_code=303)


@router.post('/ordering-tool/mappings/upsert')
async def ordering_tool_mappings_upsert(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        vendor_id = int(str(form.get('vendor_id', '')).strip())
        sku = str(form.get('sku', '')).strip()
        square_variation_id = str(form.get('square_variation_id', '')).strip() or None
        unit_cost_raw = str(form.get('unit_cost', '0')).strip() or '0'
        try:
            unit_cost = Decimal(unit_cost_raw)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError('Invalid unit cost') from exc
        pack_size = int(str(form.get('pack_size', '1')).strip() or '1')
        min_order_qty = int(str(form.get('min_order_qty', '0')).strip() or '0')
        is_default_vendor = str(form.get('is_default_vendor', 'true')).strip().lower() in {'1', 'true', 'on', 'yes'}
        active = str(form.get('active', 'true')).strip().lower() in {'1', 'true', 'on', 'yes'}
        row = upsert_vendor_sku_config(
            db,
            vendor_id=vendor_id,
            sku=sku,
            square_variation_id=square_variation_id,
            unit_cost=unit_cost,
            pack_size=pack_size,
            min_order_qty=min_order_qty,
            is_default_vendor=is_default_vendor,
            active=active,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_VENDOR_SKU_MAPPING_UPSERTED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'vendor_id': row.vendor_id, 'sku': row.sku, 'mapping_id': row.id},
    )
    db.commit()
    query = urlencode({'saved': 1, 'vendor_id': row.vendor_id})
    return RedirectResponse(f'/management/ordering-tool/mappings?{query}', status_code=303)


@router.post('/ordering-tool/mappings/import')
async def ordering_tool_mappings_import(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    upload = form.get('csv_file')
    if upload is None:
        raise HTTPException(status_code=400, detail='CSV file is required')
    try:
        content = await upload.read()
        text = content.decode('utf-8')
        result = import_vendor_sku_configs_csv(db, csv_text=text)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_VENDOR_SKU_MAPPINGS_IMPORTED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'processed': result['processed'],
            'errors': result['errors'][:20],
        },
    )
    db.commit()
    query = urlencode({'imported': result['processed'], 'errors': len(result['errors'])})
    return RedirectResponse(f'/management/ordering-tool/mappings?{query}', status_code=303)


@router.post('/ordering-tool/mappings/bulk-save')
async def ordering_tool_mappings_bulk_save(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    row_ids: set[int] = set()
    for key in form.keys():
        if '__' not in key:
            continue
        _, row_id_raw = key.rsplit('__', 1)
        if row_id_raw.isdigit():
            row_ids.add(int(row_id_raw))

    if not row_ids:
        raise HTTPException(status_code=400, detail='No mapping rows submitted')

    saved = 0
    errors: list[str] = []
    for row_id in sorted(row_ids):
        try:
            vendor_id_raw = str(form.get(f'vendor_id__{row_id}', '')).strip()
            sku = str(form.get(f'sku__{row_id}', '')).strip()
            square_variation_id = str(form.get(f'square_variation_id__{row_id}', '')).strip() or None
            unit_cost_raw = str(form.get(f'unit_cost__{row_id}', '0')).strip() or '0'
            pack_size_raw = str(form.get(f'pack_size__{row_id}', '1')).strip() or '1'
            min_order_qty_raw = str(form.get(f'min_order_qty__{row_id}', '0')).strip() or '0'
            is_default_raw = str(form.get(f'is_default_vendor__{row_id}', 'true')).strip().lower()
            active_raw = str(form.get(f'active__{row_id}', 'true')).strip().lower()

            if not vendor_id_raw.isdigit():
                raise ValueError('Invalid vendor_id')
            vendor_id = int(vendor_id_raw)
            try:
                unit_cost = Decimal(unit_cost_raw)
            except (InvalidOperation, ValueError) as exc:
                raise ValueError('Invalid unit cost') from exc

            upsert_vendor_sku_config(
                db,
                vendor_id=vendor_id,
                sku=sku,
                square_variation_id=square_variation_id,
                unit_cost=unit_cost,
                pack_size=int(pack_size_raw),
                min_order_qty=int(min_order_qty_raw),
                is_default_vendor=is_default_raw not in {'0', 'false', 'no'},
                active=active_raw not in {'0', 'false', 'no'},
            )
            saved += 1
        except Exception as exc:
            errors.append(f'Row {row_id}: {exc}')

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_VENDOR_SKU_MAPPINGS_BULK_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'saved': saved, 'errors': errors[:20]},
    )
    db.commit()
    query = urlencode({'bulk_saved': saved, 'bulk_errors': len(errors)})
    return RedirectResponse(f'/management/ordering-tool/mappings?{query}', status_code=303)


@router.post('/ordering-tool/mappings/auto-fill')
async def ordering_tool_mappings_auto_fill(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    vendor_raw = str(form.get('vendor_id', '')).strip()
    vendor_id = int(vendor_raw) if vendor_raw.isdigit() else None
    try:
        result = autofill_square_variation_ids(db, vendor_id=vendor_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_VENDOR_SKU_AUTOFILL_RUN',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'vendor_id': vendor_id, **result},
    )
    db.commit()
    q = {'autofill_updated': result['updated'], 'autofill_skipped': result['skipped']}
    if vendor_id is not None:
        q['vendor_id'] = vendor_id
    query = urlencode(q)
    return RedirectResponse(f'/management/ordering-tool/mappings?{query}', status_code=303)


@router.post('/ordering-tool/vendors/sync')
async def ordering_tool_sync_vendors(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        created, updated, deactivated = sync_vendors_from_square(db)
        mapping_sync = sync_vendor_sku_configs_from_square(db)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_VENDORS_SYNCED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'created': created, 'updated': updated, 'deactivated': deactivated, **mapping_sync},
    )
    db.commit()
    query = urlencode(
        {
            'vendors_synced': 1,
            'created': created,
            'updated': updated,
            'deactivated': deactivated,
            'map_created': mapping_sync['created'],
            'map_updated': mapping_sync['updated'],
        }
    )
    return RedirectResponse(f'/management/ordering-tool?{query}', status_code=303)


@router.post('/ordering-tool/generate')
async def ordering_tool_generate(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        vendor_ids, reorder_weeks, stock_up_weeks, history_lookback_days = parse_generation_form(form)
        created_orders, warnings = generate_purchase_orders(
            db,
            vendor_ids=vendor_ids,
            created_by_principal_id=principal.id,
            reorder_weeks=reorder_weeks,
            stock_up_weeks=stock_up_weeks,
            history_lookback_days=history_lookback_days,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDERS_GENERATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'vendor_ids': vendor_ids,
            'reorder_weeks': reorder_weeks,
            'stock_up_weeks': stock_up_weeks,
            'history_lookback_days': history_lookback_days,
            'created_order_ids': [po.id for po in created_orders],
            'warnings': warnings,
        },
    )
    db.commit()
    query_params: dict[str, object] = {'generated': len(created_orders)}
    if warnings:
        query_params['warning'] = warnings[:3]
    query = urlencode(query_params, doseq=True)
    return RedirectResponse(f'/management/ordering-tool?{query}', status_code=303)


@router.post('/ordering-tool/generate-full-stock')
async def ordering_tool_generate_full_stock(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        vendor_ids, reorder_weeks, stock_up_weeks, history_lookback_days = parse_generation_form(form)
        created_orders, warnings = generate_purchase_orders(
            db,
            vendor_ids=vendor_ids,
            created_by_principal_id=principal.id,
            reorder_weeks=reorder_weeks,
            stock_up_weeks=stock_up_weeks,
            history_lookback_days=history_lookback_days,
            include_full_stock_lines=True,
        )
    except (ValueError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDERS_FULL_STOCK_GENERATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'vendor_ids': vendor_ids,
            'reorder_weeks': reorder_weeks,
            'stock_up_weeks': stock_up_weeks,
            'history_lookback_days': history_lookback_days,
            'created_order_ids': [po.id for po in created_orders],
            'warnings': warnings,
        },
    )
    db.commit()
    query_params: dict[str, object] = {'full_stock_generated': len(created_orders)}
    if warnings:
        query_params['warning'] = warnings[:3]
    query = urlencode(query_params, doseq=True)
    return RedirectResponse(f'/management/ordering-tool?{query}', status_code=303)


@router.get('/ordering-tool/orders/{purchase_order_id}')
def ordering_tool_order_detail(
    purchase_order_id: int,
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_purchase_order_detail(db, purchase_order_id=purchase_order_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    return request.app.state.templates.TemplateResponse(
        'management_ordering_order_detail.html',
        {
            'request': request,
            'detail': detail,
            'query': request.query_params,
        },
    )


@router.get('/ordering-tool/orders/{purchase_order_id}/pdf')
def ordering_tool_order_pdf_download(
    purchase_order_id: int,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_purchase_order_detail(db, purchase_order_id=purchase_order_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    pdf_path = (detail['order'].pdf_path or '').strip()
    if not pdf_path:
        raise HTTPException(status_code=404, detail='PDF not generated for this order yet')
    abs_path = (Path(__file__).resolve().parents[2] / pdf_path).resolve()
    if not abs_path.exists() or not abs_path.is_file():
        raise HTTPException(status_code=404, detail='PDF file not found on server')
    return FileResponse(
        path=str(abs_path),
        media_type='application/pdf',
        filename=f'purchase-order-{purchase_order_id}.pdf',
    )


def _parse_order_update_form(
    form,
) -> tuple[dict[int, int], set[int], dict[int, int | None], dict[tuple[int, int], int | None], dict[tuple[int, int], int]]:
    ordered_qty_by_line_id: dict[int, int] = {}
    removed_line_ids: set[int] = set()
    manual_par_by_line_id: dict[int, int | None] = {}
    manual_par_by_line_store: dict[tuple[int, int], int | None] = {}
    allocation_qty_by_line_store: dict[tuple[int, int], int] = {}

    for key, value in form.items():
        if key.startswith('qty__'):
            line_id = int(key.split('__', 1)[1])
            raw = str(value).strip()
            ordered_qty_by_line_id[line_id] = int(raw) if raw else 0
        elif key.startswith('alloc__'):
            _, line_raw, store_raw = key.split('__', 2)
            line_id = int(line_raw)
            store_id = int(store_raw)
            raw = str(value).strip()
            allocation_qty_by_line_store[(line_id, store_id)] = int(raw) if raw else 0
        elif key.startswith('manual_par_store__'):
            _, line_raw, store_raw = key.split('__', 2)
            line_id = int(line_raw)
            store_id = int(store_raw)
            raw = str(value).strip()
            manual_par_by_line_store[(line_id, store_id)] = int(raw) if raw else None
        elif key.startswith('manual_par__'):
            line_id = int(key.split('__', 1)[1])
            raw = str(value).strip()
            manual_par_by_line_id[line_id] = int(raw) if raw else None
        elif key.startswith('remove__'):
            line_id = int(key.split('__', 1)[1])
            if str(value).strip().lower() in {'1', 'true', 'on', 'yes'}:
                removed_line_ids.add(line_id)
    return ordered_qty_by_line_id, removed_line_ids, manual_par_by_line_id, manual_par_by_line_store, allocation_qty_by_line_store


@router.post('/ordering-tool/orders/{purchase_order_id}/save')
async def ordering_tool_order_save(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        (
            ordered_qty_by_line_id,
            removed_line_ids,
            manual_par_by_line_id,
            manual_par_by_line_store,
            allocation_qty_by_line_store,
        ) = _parse_order_update_form(form)
        save_purchase_order_lines(
            db,
            purchase_order_id=purchase_order_id,
            ordered_qty_by_line_id=ordered_qty_by_line_id,
            removed_line_ids=removed_line_ids,
            manual_par_by_line_id=manual_par_by_line_id,
            manual_par_by_line_store=manual_par_by_line_store,
            allocation_qty_by_line_store=allocation_qty_by_line_store,
        )
    except (ValueError, PermissionError, RuntimeError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_DRAFT_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'purchase_order_id': purchase_order_id},
    )
    db.commit()
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?saved=1', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/add-line')
async def ordering_tool_order_add_line(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    sku = str(form.get('sku', '')).strip()
    qty_raw = str(form.get('initial_qty', '1')).strip()
    try:
        initial_qty = int(qty_raw) if qty_raw else 1
    except ValueError:
        initial_qty = 1

    try:
        line, action = add_purchase_order_line_by_sku(
            db,
            purchase_order_id=purchase_order_id,
            sku=sku,
            initial_qty=initial_qty,
        )
    except (ValueError, PermissionError, RuntimeError) as exc:
        query = urlencode({'add_error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_LINE_ADDED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'purchase_order_id': purchase_order_id,
            'line_id': line.id,
            'sku': line.sku,
            'action': action,
            'initial_qty': initial_qty,
        },
    )
    db.commit()
    query = urlencode({'added_line': 1, 'sku': line.sku or ''})
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/refresh-lines')
async def ordering_tool_order_refresh_lines(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        result = refresh_purchase_order_lines_from_catalog(
            db,
            purchase_order_id=purchase_order_id,
        )
    except (ValueError, PermissionError, RuntimeError) as exc:
        query = urlencode({'refresh_error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_LINES_REFRESHED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'purchase_order_id': purchase_order_id,
            **result,
        },
    )
    db.commit()
    query = urlencode(
        {
            'refreshed': 1,
            'refresh_updated': result['updated'],
            'refresh_scanned': result['scanned'],
            'refresh_missing': result['missing'],
        }
    )
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/submit')
async def ordering_tool_order_submit(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    try:
        (
            ordered_qty_by_line_id,
            removed_line_ids,
            manual_par_by_line_id,
            manual_par_by_line_store,
            allocation_qty_by_line_store,
        ) = _parse_order_update_form(form)
        save_purchase_order_lines(
            db,
            purchase_order_id=purchase_order_id,
            ordered_qty_by_line_id=ordered_qty_by_line_id,
            removed_line_ids=removed_line_ids,
            manual_par_by_line_id=manual_par_by_line_id,
            manual_par_by_line_store=manual_par_by_line_store,
            allocation_qty_by_line_store=allocation_qty_by_line_store,
        )
        submit_purchase_order(
            db,
            purchase_order_id=purchase_order_id,
            actor_principal_id=principal.id,
        )
    except (ValueError, PermissionError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_SUBMITTED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'purchase_order_id': purchase_order_id},
    )
    db.commit()
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?submitted=1', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/receive')
async def ordering_tool_order_receive(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        result = receive_purchase_order(
            db,
            purchase_order_id=purchase_order_id,
            retry_failed_only=False,
        )
    except (ValueError, RuntimeError) as exc:
        query = urlencode({'receive_error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)
    except Exception as exc:
        query = urlencode({'receive_error': f'Receive failed unexpectedly: {exc}'})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_SENT_TO_STORES',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'purchase_order_id': purchase_order_id,
            'store_count': result['store_count'],
            'line_count': result['line_count'],
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
            'skipped_already_synced': result['skipped_already_synced'],
        },
    )
    db.commit()
    if result['failed'] > 0:
        query = urlencode(
            {
                'receive_error': f"Square sync incomplete ({result['succeeded']} succeeded, {result['failed']} failed)",
                'receive_attempted': result['attempted'],
                'receive_succeeded': result['succeeded'],
                'receive_failed': result['failed'],
                'receive_skipped': result['skipped_already_synced'],
            }
        )
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?sent_to_stores=1', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/receive-retry-failed')
async def ordering_tool_order_receive_retry_failed(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        result = receive_purchase_order(
            db,
            purchase_order_id=purchase_order_id,
            retry_failed_only=True,
        )
    except (ValueError, RuntimeError) as exc:
        query = urlencode({'receive_error': str(exc)})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)
    except Exception as exc:
        query = urlencode({'receive_error': f'Retry failed unexpectedly: {exc}'})
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_RETRY_FAILED_SENT_TO_STORES',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'purchase_order_id': purchase_order_id,
            'store_count': result['store_count'],
            'line_count': result['line_count'],
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
            'skipped_already_synced': result['skipped_already_synced'],
            'skipped_not_failed': result['skipped_not_failed'],
        },
    )
    db.commit()
    if result['failed'] > 0:
        query = urlencode(
            {
                'receive_error': f"Retry incomplete ({result['succeeded']} succeeded, {result['failed']} failed)",
                'receive_attempted': result['attempted'],
                'receive_succeeded': result['succeeded'],
                'receive_failed': result['failed'],
                'receive_skipped': result['skipped_already_synced'],
            }
        )
        return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?{query}', status_code=303)
    return RedirectResponse(f'/management/ordering-tool/orders/{purchase_order_id}?sent_to_stores=1', status_code=303)


@router.post('/ordering-tool/orders/{purchase_order_id}/delete')
async def ordering_tool_order_delete(
    purchase_order_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        delete_draft_purchase_order(db, purchase_order_id=purchase_order_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ORDERING_PURCHASE_ORDER_DELETED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'purchase_order_id': purchase_order_id},
    )
    db.commit()
    return RedirectResponse('/management/ordering-tool?discarded=1', status_code=303)


@router.get('/daily-chore-lists')
def daily_chore_lists_page(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_sheets_for_audit(
        db,
        store_id=selected_store_id,
        from_date=from_date,
        to_date=to_date,
    )
    return request.app.state.templates.TemplateResponse(
        'management_daily_chore_audit.html',
        {
            'request': request,
            'principal': principal,
            'stores': stores,
            'rows': rows,
            'selected_store_id': selected_store_id,
            'from_date': from_raw,
            'to_date': to_raw,
            'can_delete_drafts': is_admin_role(principal.role),
        },
    )


@router.get('/daily-chore-tasks')
def daily_chore_tasks_page(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    rows = list_global_task_rows(db)
    return request.app.state.templates.TemplateResponse(
        'management_daily_chore_tasks.html',
        {
            'request': request,
            'principal': principal,
            'rows': rows,
            'sections': DAILY_CHORE_SECTION_ORDER,
            'add_task_ok': str(request.query_params.get('add_task_ok', '')).strip() == '1',
            'add_task_error': str(request.query_params.get('add_task_error', '')).strip(),
            'reorder_ok': str(request.query_params.get('reorder_ok', '')).strip() == '1',
            'reorder_error': str(request.query_params.get('reorder_error', '')).strip(),
            'delete_ok': str(request.query_params.get('delete_ok', '')).strip() == '1',
            'delete_error': str(request.query_params.get('delete_error', '')).strip(),
        },
    )


@router.post('/daily-chore-tasks/add')
async def daily_chore_tasks_add(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    section = str(form.get('section', '')).strip()
    prompt = str(form.get('prompt', '')).strip()
    order_raw = str(form.get('section_order', '')).strip()
    section_order = int(order_raw) if order_raw.isdigit() else None
    try:
        result = add_global_task(
            db,
            section=section,
            prompt=prompt,
            section_order=section_order,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'add_task_error': str(exc)})
        return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='DAILY_CHORE_TASK_TEMPLATE_ADDED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'applied_store_count': result['store_count'],
            'daily_chore_task_ids': result['task_ids'],
            'section': result['section'],
            'prompt': result['prompt'],
        },
    )
    db.commit()
    query = urlencode({'add_task_ok': '1'})
    return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)


@router.post('/daily-chore-tasks/reorder')
async def daily_chore_tasks_reorder(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    task_number_raw = str(form.get('task_number', '')).strip()
    new_number_raw = str(form.get('new_number', '')).strip()
    if not task_number_raw.isdigit() or not new_number_raw.isdigit():
        query = urlencode({'reorder_error': 'Task numbers are required'})
        return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)
    task_number = int(task_number_raw)
    new_number = int(new_number_raw)
    try:
        reorder_global_task(
            db,
            task_number=task_number,
            new_number=new_number,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'reorder_error': str(exc)})
        return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='DAILY_CHORE_TASK_TEMPLATE_REORDERED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'task_number': task_number, 'new_number': new_number},
    )
    db.commit()
    query = urlencode({'reorder_ok': '1'})
    return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)


@router.post('/daily-chore-tasks/delete')
async def daily_chore_tasks_delete(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    task_number_raw = str(form.get('task_number', '')).strip()
    if not task_number_raw.isdigit():
        query = urlencode({'delete_error': 'Task number is required'})
        return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)
    task_number = int(task_number_raw)
    try:
        result = delete_global_task(
            db,
            task_number=task_number,
        )
    except ValueError as exc:
        db.rollback()
        query = urlencode({'delete_error': str(exc)})
        return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='DAILY_CHORE_TASK_TEMPLATE_DELETED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'task_number': task_number,
            'applied_store_count': result['store_count'],
            'daily_chore_task_ids': result['task_ids'],
            'section': result['section'],
            'prompt': result['prompt'],
        },
    )
    db.commit()
    query = urlencode({'delete_ok': '1'})
    return RedirectResponse(f'/management/daily-chore-tasks?{query}', status_code=303)


@router.get('/daily-chore-lists/{sheet_id}')
def daily_chore_sheet_detail(
    sheet_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_sheet_detail_for_audit(db, sheet_id=sheet_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='DAILY_CHORE_SHEET_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'daily_chore_sheet_id': sheet_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_daily_chore_detail.html',
        {
            'request': request,
            'detail': detail,
        },
    )


@router.post('/daily-chore-lists/{sheet_id}/delete')
async def daily_chore_sheet_delete(
    sheet_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    selected_store_id_raw = str(form.get('store_id', '')).strip()
    from_raw = str(form.get('from', '')).strip()
    to_raw = str(form.get('to', '')).strip()
    query = urlencode(
        {
            key: value
            for key, value in {
                'store_id': selected_store_id_raw,
                'from': from_raw,
                'to': to_raw,
                'discarded': '1',
            }.items()
            if value
        }
    )

    try:
        sheet = delete_draft_sheet_for_management(db, sheet_id=sheet_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='DAILY_CHORE_SHEET_DRAFT_DELETED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'daily_chore_sheet_id': sheet_id, 'store_id': sheet.store_id},
    )
    db.commit()
    return RedirectResponse(
        f'/management/daily-chore-lists?{query}' if query else '/management/daily-chore-lists',
        status_code=303,
    )


@router.get('/opening-checklists')
def opening_checklists_page(
    request: Request,
    _: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()

    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_submissions(
        db,
        store_id=selected_store_id,
        from_date=from_date,
        to_date=to_date,
    )
    return request.app.state.templates.TemplateResponse(
        'management_opening_checklist_audit.html',
        {
            'request': request,
            'stores': stores,
            'rows': rows,
            'selected_store_id': selected_store_id,
            'from_date': from_raw,
            'to_date': to_raw,
        },
    )


@router.get('/opening-checklists/{submission_id}')
def opening_checklists_detail(
    submission_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_submission_detail(db, submission_id=submission_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='OPENING_CHECKLIST_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'opening_checklist_submission_id': submission_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_opening_checklist_detail.html',
        {
            'request': request,
            'detail': detail,
        },
    )


@router.get('/change-box-count')
def change_box_count_page(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_counts_for_audit(db, store_id=selected_store_id)
    return request.app.state.templates.TemplateResponse(
        'management_change_box_count_audit.html',
        {
            'request': request,
            'principal': principal,
            'stores': stores,
            'rows': rows,
            'selected_store_id': selected_store_id,
        },
    )


@router.get('/change-box-count/{count_id}')
def change_box_count_detail(
    count_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_count_detail(db, count_id=count_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CHANGE_BOX_COUNT_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'change_box_count_id': count_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_change_box_count_detail.html',
        {
            'request': request,
            'principal': principal,
            'detail': detail,
        },
    )


@router.post('/change-box-count/{count_id}/delete')
async def change_box_count_delete(
    count_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        deleted = delete_change_box_count(db, count_id=count_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CHANGE_BOX_COUNT_DELETED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'change_box_count_id': deleted['id'],
            'store_id': deleted['store_id'],
            'status': deleted['status'],
        },
    )
    db.commit()
    return RedirectResponse('/management/change-box-count', status_code=303)


@router.get('/change-forms')
def change_forms_page(
    request: Request,
    _: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_change_forms(db, store_id=selected_store_id)
    return request.app.state.templates.TemplateResponse(
        'management_change_forms.html',
        {
            'request': request,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'rows': rows,
        },
    )


@router.get('/change-forms/{submission_id}')
def change_form_detail(
    submission_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_change_form_detail(db, submission_id=submission_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CHANGE_FORM_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'change_form_submission_id': submission_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_change_form_detail.html',
        {
            'request': request,
            'detail': detail,
        },
    )


@router.get('/change-box-audit')
def change_box_audit_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else (stores[0].id if stores else None)
    inventory = get_inventory_state(db, store_id=selected_store_id) if selected_store_id else {'target_amount': 0, 'total_amount': 0, 'lines': []}
    return request.app.state.templates.TemplateResponse(
        'management_change_box_audit.html',
        {
            'request': request,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'inventory': inventory,
            'denoms': DENOMS,
            'roll_sizes': ROLL_SIZES_BY_CODE,
        },
    )


@router.get('/exchange-return-forms')
def exchange_return_forms_page(
    request: Request,
    _: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_exchange_return_forms(
        db,
        store_id=selected_store_id,
        from_date=from_date,
        to_date=to_date,
    )
    return request.app.state.templates.TemplateResponse(
        'management_exchange_return_forms.html',
        {
            'request': request,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'from_date': from_raw,
            'to_date': to_raw,
            'rows': rows,
        },
    )


@router.get('/exchange-return-forms/{form_id}')
def exchange_return_form_detail(
    form_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_exchange_return_form_detail(db, form_id=form_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='EXCHANGE_RETURN_FORM_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'exchange_return_form_id': form_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_exchange_return_form_detail.html',
        {
            'request': request,
            'detail': detail,
        },
    )


@router.post('/change-box-audit/{store_id}/submit')
async def change_box_audit_submit(
    store_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    auditor_name = str(form.get('auditor_name', '')).strip()
    target_amount_raw = str(form.get('target_amount', '0')).strip()
    quantities_by_code: dict[str, int] = {}
    for denom in DENOMS:
        code = denom['code']
        rolls_raw = str(form.get(f'qty_rolls__{code}', '0')).strip()
        loose_raw = str(form.get(f'qty_loose__{code}', '0')).strip()
        rolls = int(rolls_raw) if rolls_raw else 0
        loose = int(loose_raw) if loose_raw else 0
        if code in ROLL_SIZES_BY_CODE:
            quantities_by_code[code] = (rolls * ROLL_SIZES_BY_CODE[code]) + loose
        else:
            quantities_by_code[code] = loose

    try:
        target_amount = Decimal(target_amount_raw or '0')
    except InvalidOperation as exc:
        raise HTTPException(status_code=400, detail='Invalid target amount') from exc

    try:
        audit = submit_inventory_audit(
            db,
            store_id=store_id,
            principal_id=principal.id,
            auditor_name=auditor_name,
            target_amount=target_amount,
            quantities_by_code=quantities_by_code,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CHANGE_BOX_AUDIT_SUBMITTED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'change_box_audit_submission_id': audit.id, 'store_id': store_id},
    )
    db.commit()
    return RedirectResponse(f'/management/change-box-audit?store_id={store_id}', status_code=303)


@router.get('/master-safe-audit')
def master_safe_audit_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    inventory = get_master_safe_inventory_state(db)
    return request.app.state.templates.TemplateResponse(
        'management_master_safe_audit.html',
        {
            'request': request,
            'inventory': inventory,
            'roll_sizes': ROLL_SIZES_BY_CODE,
        },
    )


@router.post('/master-safe-audit/submit')
async def master_safe_audit_submit(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    auditor_name = str(form.get('auditor_name', '')).strip()
    target_amount_raw = str(form.get('target_amount', '0')).strip()
    quantities_by_code: dict[str, int] = {}
    for denom in DENOMS:
        code = denom['code']
        rolls_raw = str(form.get(f'qty_rolls__{code}', '0')).strip()
        loose_raw = str(form.get(f'qty_loose__{code}', '0')).strip()
        rolls = int(rolls_raw) if rolls_raw else 0
        loose = int(loose_raw) if loose_raw else 0
        if code in ROLL_SIZES_BY_CODE:
            quantities_by_code[code] = (rolls * ROLL_SIZES_BY_CODE[code]) + loose
        else:
            quantities_by_code[code] = loose

    try:
        target_amount = Decimal(target_amount_raw or '0')
    except InvalidOperation as exc:
        raise HTTPException(status_code=400, detail='Invalid target amount') from exc

    try:
        audit = submit_master_safe_audit(
            db,
            principal_id=principal.id,
            auditor_name=auditor_name,
            target_amount=target_amount,
            quantities_by_code=quantities_by_code,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='MASTER_SAFE_AUDIT_SUBMITTED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'master_safe_audit_submission_id': audit.id},
    )
    db.commit()
    return RedirectResponse('/management/master-safe-audit', status_code=303)


@router.get('/non-sellable-stock-take')
def non_sellable_stock_take_page(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    stock_takes = list_stock_takes_for_audit(db, store_id=selected_store_id, include_draft=True)
    items = list_non_sellable_items(db, include_inactive=True)
    return request.app.state.templates.TemplateResponse(
        'management_non_sellable_stock_take.html',
        {
            'request': request,
            'principal': principal,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'stock_takes': stock_takes,
            'items': items,
            'can_manage_items': is_admin_role(principal.role),
        },
    )


@router.get('/non-sellable-stock-take/{stock_take_id}')
def non_sellable_stock_take_detail(
    stock_take_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    try:
        detail = get_stock_take_detail(db, stock_take_id=stock_take_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='NON_SELLABLE_STOCK_TAKE_VIEWED_AUDIT',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'non_sellable_stock_take_id': stock_take_id},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_non_sellable_stock_take_detail.html',
        {
            'request': request,
            'detail': detail,
            'principal': principal,
        },
    )


@router.post('/non-sellable-stock-take/{stock_take_id}/unlock')
async def non_sellable_stock_take_unlock(
    stock_take_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        take = unlock_stock_take(db, stock_take_id=stock_take_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='NON_SELLABLE_STOCK_TAKE_UNLOCKED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'non_sellable_stock_take_id': take.id},
    )
    db.commit()
    return RedirectResponse(f'/management/non-sellable-stock-take/{take.id}', status_code=303)


@router.post('/non-sellable-stock-take/items/create')
async def non_sellable_item_create(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    name = str(form.get('name', '')).strip()
    try:
        item = add_non_sellable_item(db, name=name, created_by_principal_id=principal.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='NON_SELLABLE_ITEM_CREATED_OR_REACTIVATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'item_id': item.id, 'name': item.name},
    )
    db.commit()
    return RedirectResponse('/management/non-sellable-stock-take', status_code=303)


@router.post('/non-sellable-stock-take/items/{item_id}/deactivate')
async def non_sellable_item_deactivate(
    item_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        item = deactivate_non_sellable_item(db, item_id=item_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='NON_SELLABLE_ITEM_DEACTIVATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'item_id': item.id, 'name': item.name},
    )
    db.commit()
    return RedirectResponse('/management/non-sellable-stock-take', status_code=303)


@router.get('/customer-requests')
def customer_requests_page(
    request: Request,
    _: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    submissions = list_customer_request_submissions(
        db,
        store_id=selected_store_id,
        from_date=from_date,
        to_date=to_date,
    )
    items = list_customer_request_items_for_management(db)
    return request.app.state.templates.TemplateResponse(
        'management_customer_requests.html',
        {
            'request': request,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'from_date': from_raw,
            'to_date': to_raw,
            'submissions': submissions,
            'items': items,
        },
    )


@router.post('/customer-requests/items/create')
async def customer_requests_item_create(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    name = str(form.get('name', '')).strip()
    try:
        item = add_customer_request_item(db, name=name, principal_id=principal.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CUSTOMER_REQUEST_ITEM_CREATED_OR_REACTIVATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'item_id': item.id, 'name': item.name},
    )
    db.commit()
    return RedirectResponse('/management/customer-requests', status_code=303)


@router.post('/customer-requests/items/{item_id}/count')
async def customer_requests_item_set_count(
    item_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    count_raw = str(form.get('request_count', '0')).strip()
    try:
        request_count = int(count_raw)
        item = set_customer_request_item_count(db, item_id=item_id, request_count=request_count)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='CUSTOMER_REQUEST_ITEM_COUNT_UPDATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'item_id': item.id, 'request_count': item.request_count},
    )
    db.commit()
    return RedirectResponse('/management/customer-requests', status_code=303)


@router.get('/audit-queue')
def audit_queue_page(request: Request, _: Principal = Depends(management_access)):
    return _render_placeholder(request, 'Audit Queue')


@router.get('/reports')
def reports_page(request: Request, _: Principal = Depends(management_access)):
    return request.app.state.templates.TemplateResponse(
        'management_reports.html',
        {
            'request': request,
        },
    )


@router.get('/reports/count-square-sync')
def count_square_sync_report_page(
    request: Request,
    _: Principal = Depends(require_role(Role.ADMIN)),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()
    session_id_raw = request.query_params.get('session_id', '').strip()
    sync_scope_raw = request.query_params.get('sync_scope', '').strip().lower()

    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    selected_session_id = int(session_id_raw) if session_id_raw.isdigit() else None
    sync_scope = 'recount' if sync_scope_raw == 'recount' else 'all'
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    rows = list_count_square_sync_report_rows(
        db,
        store_id=selected_store_id,
        from_date=from_date,
        to_date=to_date,
        session_id=selected_session_id,
        recount_only=sync_scope == 'recount',
        limit=1000,
    )
    return request.app.state.templates.TemplateResponse(
        'management_count_square_sync_report.html',
        {
            'request': request,
            'stores': stores,
            'rows': rows,
            'selected_store_id': selected_store_id,
            'selected_session_id': selected_session_id,
            'from_date': from_raw,
            'to_date': to_raw,
            'sync_scope': sync_scope,
        },
    )


@router.get('/reports/recount-changes')
def recount_change_report_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = request.query_params.get('store_id', '').strip()
    from_raw = request.query_params.get('from', '').strip()
    to_raw = request.query_params.get('to', '').strip()

    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        from_date = date.fromisoformat(from_raw) if from_raw else None
        to_date = date.fromisoformat(to_raw) if to_raw else None
    except ValueError as exc:
        raise HTTPException(status_code=400, detail='Invalid date filter') from exc

    query = (
        select(CountSession.id, CountSession.store_id, CountSession.submitted_at, Store.name)
        .join(Store, Store.id == CountSession.store_id)
        .where(
            CountSession.status == SessionStatus.SUBMITTED,
            CountSession.includes_recount.is_(True),
            CountSession.stable_variance.is_(True),
            CountSession.submitted_at.is_not(None),
        )
        .order_by(CountSession.submitted_at.desc(), CountSession.id.desc())
    )
    if selected_store_id is not None:
        query = query.where(CountSession.store_id == selected_store_id)
    if from_date is not None:
        query = query.where(CountSession.submitted_at >= datetime.combine(from_date, datetime.min.time()))
    if to_date is not None:
        query = query.where(CountSession.submitted_at < datetime.combine(to_date + timedelta(days=1), datetime.min.time()))

    sessions = db.execute(query.limit(300)).all()
    rows: list[dict] = []
    for session in sessions:
        variance_rows = get_management_variance_lines(db, session_id=int(session.id))
        for line in variance_rows:
            if str(line.get('section_type') or '').upper() != 'RECOUNT':
                continue
            if not bool(line.get('recount_closed_out')):
                continue
            variance = Decimal(str(line.get('variance') or '0'))
            if variance == 0:
                continue
            expected = Decimal(str(line.get('expected_on_hand') or '0'))
            counted = Decimal(str(line.get('counted_qty') or '0'))
            rows.append(
                {
                    'session_id': int(session.id),
                    'store_id': int(session.store_id),
                    'store_name': session.name,
                    'submitted_at': session.submitted_at,
                    'variation_id': str(line.get('variation_id') or ''),
                    'sku': line.get('sku'),
                    'item_name': line.get('item_name'),
                    'variation_name': line.get('variation_name'),
                    'expected_on_hand': expected,
                    'counted_qty': counted,
                    'delta': variance,
                }
            )

    stores = db.execute(select(Store.id, Store.name).where(Store.active.is_(True)).order_by(Store.name.asc())).all()
    return request.app.state.templates.TemplateResponse(
        'management_recount_change_report.html',
        {
            'request': request,
            'stores': stores,
            'rows': rows,
            'selected_store_id': selected_store_id,
            'from_date': from_raw,
            'to_date': to_raw,
        },
    )


@router.get('/reports/cogs')
def reports_cogs_page(
    request: Request,
    _: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    query = request.query_params
    start_raw = str(query.get('start_date', '')).strip()
    end_raw = str(query.get('end_date', '')).strip()

    today = date.today()
    default_start = (today - timedelta(days=6)).isoformat()
    default_end = today.isoformat()

    report = None
    error = None
    if start_raw or end_raw:
        if not start_raw or not end_raw:
            error = 'Both start date and end date are required.'
        else:
            try:
                start_date = date.fromisoformat(start_raw)
                end_date = date.fromisoformat(end_raw)
                report = build_cogs_report(db, start_date=start_date, end_date=end_date)
            except ValueError as exc:
                error = str(exc)
            except RuntimeError as exc:
                error = str(exc)

    return request.app.state.templates.TemplateResponse(
        'management_cogs_report.html',
        {
            'request': request,
            'start_date': start_raw or default_start,
            'end_date': end_raw or default_end,
            'report': report,
            'error': error,
        },
    )


@router.get('/reports/stock-value-on-hand')
def reports_stock_value_on_hand_page(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = str(request.query_params.get('store_id', '')).strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    stores = db.execute(
        select(Store.id, Store.name)
        .where(Store.active.is_(True), Store.square_location_id.is_not(None))
        .order_by(Store.name.asc())
    ).all()

    report = None
    error = None
    try:
        report = build_stock_value_on_hand_report(db, store_id=selected_store_id, top_n_items=200)
    except RuntimeError as exc:
        error = str(exc)

    return request.app.state.templates.TemplateResponse(
        'management_stock_value_on_hand.html',
        {
            'request': request,
            'stores': stores,
            'selected_store_id': selected_store_id,
            'report': report,
            'error': error,
        },
    )


@router.get('/reports/stock-value-on-hand/export.csv')
def reports_stock_value_on_hand_export_csv(
    request: Request,
    _: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
):
    selected_store_id_raw = str(request.query_params.get('store_id', '')).strip()
    selected_store_id = int(selected_store_id_raw) if selected_store_id_raw.isdigit() else None
    try:
        report = build_stock_value_on_hand_report(db, store_id=selected_store_id, top_n_items=None)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(['Stock Value On Hand'])
    writer.writerow(['As Of (UTC)', report.as_of_utc.isoformat()])
    writer.writerow(['Store Filter', selected_store_id or 'ALL_ACTIVE_STORES'])
    writer.writerow([])

    writer.writerow(['Summary Metric', 'Value'])
    writer.writerow(['Active Stores', report.active_store_count])
    writer.writerow(['Tracked Variations', report.tracked_variation_count])
    writer.writerow(['In-stock Variations', report.in_stock_variation_count])
    writer.writerow(['Total Units On Hand', f'{report.total_units_on_hand:.3f}'])
    writer.writerow(['Total Cost Value', f'{report.total_cost_value:.2f}'])
    writer.writerow(['Total Retail Value', f'{report.total_retail_value:.2f}'])
    writer.writerow(['In-stock Variations Missing Cost', report.missing_cost_variation_count])
    writer.writerow(['In-stock Variations Missing Price', report.missing_price_variation_count])
    writer.writerow([])

    writer.writerow(['By Store'])
    writer.writerow(['Store ID', 'Store Name', 'Units On Hand', 'Cost Value', 'Retail Value'])
    for row in report.store_rows:
        writer.writerow(
            [
                row.store_id,
                row.store_name,
                f'{row.total_units_on_hand:.3f}',
                f'{row.total_cost_value:.2f}',
                f'{row.total_retail_value:.2f}',
            ]
        )
    writer.writerow([])

    writer.writerow(['Item Detail (Sorted by Extended Retail Value)'])
    writer.writerow(
        [
            'Variation ID',
            'Item Name',
            'Variation Name',
            'On Hand Qty',
            'Unit Cost',
            'Unit Price',
            'Extended Cost',
            'Extended Retail',
        ]
    )
    for row in report.top_item_rows:
        writer.writerow(
            [
                row.variation_id,
                row.item_name,
                row.variation_name,
                f'{row.on_hand_qty:.3f}',
                f'{row.unit_cost:.2f}' if row.unit_cost is not None else '',
                f'{row.unit_price:.2f}' if row.unit_price is not None else '',
                f'{row.extended_cost_value:.2f}',
                f'{row.extended_retail_value:.2f}',
            ]
        )

    timestamp = datetime.now().astimezone().strftime('%Y%m%d_%H%M%S')
    scope = f'store_{selected_store_id}' if selected_store_id is not None else 'all_stores'
    filename = f"stock_value_on_hand_{_safe_excel_filename_part(scope)}_{timestamp}.csv"
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(
        iter([sio.getvalue()]),
        media_type='text/csv',
        headers=headers,
    )


@router.get('/users')
def users_page(
    request: Request,
    _: Principal = Depends(users_access),
    db: Session = Depends(get_db),
):
    users = list_management_users(db)
    return request.app.state.templates.TemplateResponse(
        'management_users.html',
        {
            'request': request,
            'users': users,
        },
    )


@router.post('/users/create')
async def create_user(
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    username = str(form.get('username', '')).strip()
    password = str(form.get('password', ''))
    role = str(form.get('role', 'LEAD')).strip().upper()
    try:
        created = create_management_user(
            db,
            actor=principal,
            username=username,
            password=password,
            role=role,
        )
    except (ValueError, PermissionError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='MANAGEMENT_USER_CREATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'principal_id': created.id, 'username': created.username, 'role': created.role.value},
    )
    db.commit()
    return RedirectResponse('/management/users', status_code=303)


@router.post('/users/{target_principal_id}/status')
async def set_user_status(
    target_principal_id: int,
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    active = str(form.get('active', '')).strip().lower() in {'1', 'true', 'yes', 'on'}
    try:
        updated = set_management_user_active(
            db,
            actor=principal,
            target_principal_id=target_principal_id,
            active=active,
        )
    except (ValueError, PermissionError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='MANAGEMENT_USER_STATUS_UPDATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'principal_id': updated.id, 'active': updated.active},
    )
    db.commit()
    return RedirectResponse('/management/users', status_code=303)


@router.post('/users/{target_principal_id}/password')
async def set_user_password(
    target_principal_id: int,
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    new_password = str(form.get('new_password', ''))
    try:
        updated = reset_management_user_password(
            db,
            actor=principal,
            target_principal_id=target_principal_id,
            new_password=new_password,
        )
    except (ValueError, PermissionError) as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='MANAGEMENT_USER_PASSWORD_RESET',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'principal_id': updated.id},
    )
    db.commit()
    return RedirectResponse('/management/users', status_code=303)


@router.get('/access-controls')
def access_controls_page(
    request: Request,
    _: Principal = Depends(users_access),
    db: Session = Depends(get_db),
):
    detail = list_access_control_settings(db)
    return request.app.state.templates.TemplateResponse(
        'management_access_controls.html',
        {
            'request': request,
            'detail': detail,
            'saved': str(request.query_params.get('saved', '')).strip() == '1',
        },
    )


@router.post('/access-controls/roles/save')
async def access_controls_save_roles(
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    allowed_map: dict[tuple[str, str], bool] = {}
    defs = permission_defs()
    roles = ['ADMIN', 'MANAGER', 'LEAD', 'STORE']
    for role in roles:
        for permission in defs:
            key = f'role_perm__{role}__{permission.key}'
            allowed_map[(role, permission.key)] = str(form.get(key, '')).strip().lower() in {'1', 'true', 'on', 'yes'}
    save_role_permission_overrides(
        db,
        actor_principal_id=principal.id,
        allowed_map=allowed_map,
    )
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ACCESS_CONTROLS_ROLE_OVERRIDES_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'role_count': len(roles), 'permission_count': len(defs)},
    )
    db.commit()
    return RedirectResponse('/management/access-controls?saved=1', status_code=303)


@router.get('/access-controls/roles/{role}/categories')
def access_controls_role_categories_page(
    role: str,
    request: Request,
    _: Principal = Depends(users_access),
    db: Session = Depends(get_db),
):
    try:
        detail = list_role_dashboard_category_access(db, role=role)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return request.app.state.templates.TemplateResponse(
        'management_access_role_categories.html',
        {
            'request': request,
            'detail': detail,
            'saved': str(request.query_params.get('saved', '')).strip() == '1',
        },
    )


@router.post('/access-controls/roles/{role}/categories/save')
async def access_controls_role_categories_save(
    role: str,
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    allowed_by_category_id: dict[int, bool] = {}
    for key in form.keys():
        key_text = str(key)
        if not key_text.startswith('category__'):
            continue
        category_id_raw = key_text.split('category__', 1)[1]
        if not category_id_raw.isdigit():
            continue
        category_id = int(category_id_raw)
        allowed_by_category_id[category_id] = str(form.get(key, '')).strip().upper() == 'YES'
    try:
        save_role_dashboard_category_access(
            db,
            role=role,
            actor_principal_id=principal.id,
            allowed_by_category_id=allowed_by_category_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ACCESS_CONTROLS_ROLE_CATEGORY_ACCESS_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'role': role.upper(), 'category_count': len(allowed_by_category_id)},
    )
    db.commit()
    return RedirectResponse(f'/management/access-controls/roles/{role.upper()}/categories?saved=1', status_code=303)


@router.post('/access-controls/principals/save')
async def access_controls_save_principals(
    request: Request,
    principal: Principal = Depends(users_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    defs = permission_defs()
    override_map: dict[tuple[int, str], str] = {}
    custom_role_labels: dict[int, str] = {}
    principal_ids = [int(value) for value in form.getlist('principal_id') if str(value).strip().isdigit()]
    for principal_id in principal_ids:
        label_key = f'custom_role_label__{principal_id}'
        custom_role_labels[principal_id] = str(form.get(label_key, '')).strip()
        for permission in defs:
            key = f'principal_perm__{principal_id}__{permission.key}'
            state = str(form.get(key, 'DEFAULT')).strip().upper()
            if state not in {'ALLOW', 'DENY', 'DEFAULT'}:
                state = 'DEFAULT'
            override_map[(principal_id, permission.key)] = state
    save_principal_permission_overrides(
        db,
        actor_principal_id=principal.id,
        override_map=override_map,
        custom_role_labels=custom_role_labels,
    )
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='ACCESS_CONTROLS_PRINCIPAL_OVERRIDES_SAVED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'principal_count': len(principal_ids), 'permission_count': len(defs)},
    )
    db.commit()
    return RedirectResponse('/management/access-controls?saved=1', status_code=303)


@router.get('/sessions')
def list_sessions(
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    rows = db.execute(
        select(
            CountSession.id,
            CountSession.employee_name,
            CountSession.status,
            CountSession.includes_recount,
            CountSession.source_forced_count_id,
            CountSession.created_at,
            CountSession.submitted_at,
            Store.name.label('store_name'),
            CountGroup.name.label('group_name'),
            Campaign.category_filter.label('campaign_category_filter'),
            Campaign.label.label('campaign_label'),
        )
        .join(Store, Store.id == CountSession.store_id)
        .join(Campaign, Campaign.id == CountSession.campaign_id)
        .outerjoin(CountGroup, CountGroup.id == CountSession.count_group_id)
        .order_by(CountSession.created_at.desc())
    ).all()

    return request.app.state.templates.TemplateResponse(
        'management_sessions.html',
        {
            'request': request,
            'principal': principal,
            'rows': rows,
        },
    )


@router.post('/sessions/delete')
async def delete_sessions(
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    session_ids = [int(v) for v in form.getlist('session_ids')]
    deleted_count = purge_count_sessions(db, session_ids=session_ids)
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSIONS_PURGED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'requested_ids': session_ids, 'deleted_count': deleted_count},
    )
    db.commit()
    return RedirectResponse('/management/sessions', status_code=303)


@router.get('/groups')
def groups_page(
    request: Request,
    _: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
):
    params = request.query_params
    data = group_management_data(db)
    store_rotation_rows = list_stores_with_rotation(db)
    store_login_rows = list_store_login_rows(db)
    return request.app.state.templates.TemplateResponse(
        'management_groups.html',
        {
            'request': request,
            'groups': data['groups'],
            'ungrouped_campaigns': data['ungrouped_campaigns'],
            'campaign_rows': data['campaign_rows'],
            'store_rotation_rows': store_rotation_rows,
            'store_login_rows': store_login_rows,
            'sync_summary': {
                'created': params.get('created'),
                'updated': params.get('updated'),
                'deactivated': params.get('deactivated'),
            }
            if params.get('created') is not None
            else None,
        },
    )


@router.get('/groups/audit-count-groups')
def audit_count_groups_page(
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
):
    try:
        audit_data = run_count_group_coverage_audit(db)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_GROUP_AUDIT_RUN',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'variation_count': audit_data['summary']['variation_count'],
            'uncovered_variation_count': audit_data['summary']['uncovered_variation_count'],
            'overlap_variation_count': audit_data['summary']['overlap_variation_count'],
        },
    )
    db.commit()

    return request.app.state.templates.TemplateResponse(
        'management_group_coverage_audit.html',
        {
            'request': request,
            'summary': audit_data['summary'],
            'group_rows': audit_data['group_rows'],
            'category_rows': audit_data['category_rows'],
            'ungrouped_campaign_rows': audit_data['ungrouped_campaign_rows'],
            'uncovered_rows': audit_data['uncovered_rows'],
            'uncovered_remaining_count': audit_data['uncovered_remaining_count'],
            'overlap_rows': audit_data['overlap_rows'],
            'overlap_remaining_count': audit_data['overlap_remaining_count'],
        },
    )


@router.post('/groups/create')
async def create_group(
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    name = str(form.get('name', '')).strip()
    campaign_ids = [int(v) for v in form.getlist('campaign_ids')]

    try:
        group = create_count_group(db, name=name, campaign_ids=campaign_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_GROUP_CREATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'group_id': group.id, 'name': name, 'campaign_ids': campaign_ids},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/stores/{store_id}/credentials')
async def update_store_credentials(
    store_id: int,
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    username = str(form.get('username', '')).strip()
    password = str(form.get('password', '')).strip()

    try:
        updated_principal, created = upsert_store_login_credentials(
            db,
            store_id=store_id,
            username=username,
            new_password=password if password else None,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='STORE_LOGIN_CREDENTIALS_UPDATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'store_id': store_id,
            'principal_id': updated_principal.id,
            'username': updated_principal.username,
            'password_changed': bool(password),
            'created': created,
        },
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/password/reset')
async def reset_password(
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    current_password = str(form.get('current_password', ''))
    new_password = str(form.get('new_password', ''))
    confirm_password = str(form.get('confirm_password', ''))

    try:
        reset_manager_password(
            db,
            manager_principal_id=principal.id,
            current_password=current_password,
            new_password=new_password,
            confirm_password=confirm_password,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='MANAGER_PASSWORD_CHANGED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/groups/{group_id}/update')
async def update_group(
    group_id: int,
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    name = str(form.get('name', '')).strip()
    campaign_ids = [int(v) for v in form.getlist('campaign_ids')]

    try:
        group = update_count_group(db, group_id=group_id, name=name, campaign_ids=campaign_ids)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_GROUP_UPDATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'group_id': group.id, 'name': group.name, 'campaign_ids': campaign_ids},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/groups/{group_id}/delete')
async def delete_group(
    group_id: int,
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        group = deactivate_count_group(db, group_id=group_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_GROUP_DEACTIVATED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'group_id': group.id, 'name': group.name},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/groups/renumber')
async def renumber_groups(
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    changed = renumber_count_group_positions(db)
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_GROUP_POSITIONS_RENUMBERED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'changed_rows': changed},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.post('/groups/sync-campaigns')
async def sync_campaigns_from_square(
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    min_items = int(str(form.get('min_items', '1')).strip() or '1')
    if min_items < 1:
        min_items = 1
    deactivate_missing = str(form.get('deactivate_missing', '')).lower() in {'1', 'true', 'on', 'yes'}

    try:
        created, updated, deactivated = sync_campaigns(
            min_items=min_items,
            deactivate_missing=deactivate_missing,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='SQUARE_CAMPAIGNS_SYNCED',
        session_id=None,
        ip=get_client_ip(request),
        metadata={
            'min_items': min_items,
            'deactivate_missing': deactivate_missing,
            'created': created,
            'updated': updated,
            'deactivated': deactivated,
        },
    )
    db.commit()

    query = urlencode({'created': created, 'updated': updated, 'deactivated': deactivated})
    return RedirectResponse(f'/management/groups?{query}', status_code=303)


@router.post('/stores/{store_id}/set-next-group')
async def set_next_group(
    store_id: int,
    request: Request,
    principal: Principal = Depends(groups_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    form = await request.form()
    group_id = int(form.get('group_id'))

    try:
        rotation = set_store_next_group(db, store_id=store_id, group_id=group_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='STORE_NEXT_GROUP_SET',
        session_id=None,
        ip=get_client_ip(request),
        metadata={'store_id': store_id, 'group_id': group_id, 'next_group_id': rotation.next_group_id},
    )
    db.commit()
    return RedirectResponse('/management/groups', status_code=303)


@router.get('/sessions/{session_id}')
def view_session(
    session_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    session_row = db.execute(
        select(
            CountSession.id,
            CountSession.store_id,
            CountSession.campaign_id,
            CountSession.count_group_id,
            CountSession.employee_name,
            CountSession.status,
            CountSession.stable_variance,
            CountSession.includes_recount,
            CountSession.source_forced_count_id,
            CountSession.created_at,
            CountSession.submitted_at,
            Store.name.label('store_name'),
            CountGroup.name.label('group_name'),
            Campaign.category_filter.label('campaign_category_filter'),
            Campaign.label.label('campaign_label'),
        )
        .join(Store, Store.id == CountSession.store_id)
        .join(Campaign, Campaign.id == CountSession.campaign_id)
        .outerjoin(CountGroup, CountGroup.id == CountSession.count_group_id)
        .where(CountSession.id == session_id)
    ).one_or_none()
    if not session_row:
        return RedirectResponse('/management/sessions', status_code=303)

    variance_rows = get_management_variance_lines(db, session_id=session_id)
    for row in variance_rows:
        if str(row.get('section_type') or '').upper() != 'RECOUNT':
            row['previous_recount_variance'] = None
            row['recount_match'] = None
            continue
        prior = row.get('previous_recount_variance')
        row['previous_recount_variance'] = prior
        if prior is None:
            row['recount_match'] = None
        else:
            row['recount_match'] = Decimal(str(row.get('variance') or '0')) == prior
    no_variance = all(row['variance'] == 0 for row in variance_rows)
    is_submitted = session_row.status.value == 'SUBMITTED'
    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSION_VIEWED_MANAGER',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={},
    )
    db.commit()
    return request.app.state.templates.TemplateResponse(
        'management_session_detail.html',
        {
            'request': request,
            'principal': principal,
            'session_row': session_row,
            'variance_rows': variance_rows,
            'no_variance': no_variance,
            'is_submitted': is_submitted,
            'can_force_recount': is_admin_role(principal.role),
            'can_push_to_square': principal.role == Role.ADMIN and is_submitted,
            'push_square_attempted': request.query_params.get('push_square_attempted'),
            'push_square_succeeded': request.query_params.get('push_square_succeeded'),
            'push_square_failed': request.query_params.get('push_square_failed'),
            'push_square_error': request.query_params.get('push_square_error'),
            'push_recount_square_attempted': request.query_params.get('push_recount_square_attempted'),
            'push_recount_square_succeeded': request.query_params.get('push_recount_square_succeeded'),
            'push_recount_square_failed': request.query_params.get('push_recount_square_failed'),
            'push_recount_square_error': request.query_params.get('push_recount_square_error'),
        },
    )


@router.post('/sessions/{session_id}/push-to-square')
def push_session_to_square(
    session_id: int,
    request: Request,
    principal: Principal = Depends(require_role(Role.ADMIN)),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        result = push_session_variance_to_square(
            db,
            session_id=session_id,
        )
    except (ValueError, RuntimeError) as exc:
        query = urlencode({'push_square_error': str(exc)})
        return RedirectResponse(f'/management/sessions/{session_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSION_PUSHED_TO_SQUARE',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={
            'store_id': result['store_id'],
            'location_id': result['location_id'],
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
        },
    )
    db.commit()
    query = urlencode(
        {
            'push_square_attempted': str(result['attempted']),
            'push_square_succeeded': str(result['succeeded']),
            'push_square_failed': str(result['failed']),
        }
    )
    return RedirectResponse(f'/management/sessions/{session_id}?{query}', status_code=303)


@router.post('/sessions/{session_id}/push-recount-to-square')
def push_session_recount_to_square(
    session_id: int,
    request: Request,
    principal: Principal = Depends(require_role(Role.ADMIN)),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        result = push_session_recount_variance_to_square(
            db,
            session_id=session_id,
        )
    except (ValueError, RuntimeError) as exc:
        query = urlencode({'push_recount_square_error': str(exc)})
        return RedirectResponse(f'/management/sessions/{session_id}?{query}', status_code=303)

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSION_RECOUNT_PUSHED_TO_SQUARE',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={
            'store_id': result['store_id'],
            'location_id': result['location_id'],
            'attempted': result['attempted'],
            'succeeded': result['succeeded'],
            'failed': result['failed'],
        },
    )
    db.commit()
    query = urlencode(
        {
            'push_recount_square_attempted': str(result['attempted']),
            'push_recount_square_succeeded': str(result['succeeded']),
            'push_recount_square_failed': str(result['failed']),
        }
    )
    return RedirectResponse(f'/management/sessions/{session_id}?{query}', status_code=303)


@router.post('/sessions/{session_id}/force-recount')
def force_recount(
    session_id: int,
    request: Request,
    principal: Principal = Depends(admin_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    session_row = db.execute(select(CountSession).where(CountSession.id == session_id)).scalar_one_or_none()
    if not session_row:
        raise HTTPException(status_code=404, detail='Session not found')

    forced = create_forced_count(
        db,
        manager_principal_id=principal.id,
        store_id=session_row.store_id,
        group_id=session_row.count_group_id,
        campaign_id=session_row.campaign_id,
        reason='Manager forced recount from submitted session',
        source_session_id=session_id,
    )

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='FORCE_RECOUNT_CREATED',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={
            'forced_count_id': forced.id,
            'campaign_id': session_row.campaign_id,
            'group_id': session_row.count_group_id,
            'store_id': session_row.store_id,
        },
    )
    db.commit()
    return RedirectResponse('/management/sessions', status_code=303)


@router.post('/sessions/{session_id}/unlock')
def unlock(
    session_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
    _: None = Depends(verify_csrf),
):
    try:
        count_session = unlock_session(db, principal=principal, session_id=session_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSION_UNLOCKED',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={'previous_status': 'SUBMITTED', 'new_status': 'DRAFT'},
    )
    db.commit()
    return RedirectResponse(f'/management/sessions/{count_session.id}', status_code=303)


@router.get('/sessions/{session_id}/export.csv')
def export_csv(
    session_id: int,
    request: Request,
    principal: Principal = Depends(management_access),
    db: Session = Depends(get_db),
):
    variance_rows = get_management_variance_lines(db, session_id=session_id)

    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(['Section', 'SKU', 'Item Name', 'Variation', 'Expected On Hand', 'Counted Qty', 'Variance'])
    for row in variance_rows:
        writer.writerow(
            [
                row['section_type'],
                row['sku'] or '',
                row['item_name'],
                row['variation_name'],
                row['expected_on_hand'],
                row['counted_qty'],
                row['variance'],
            ]
        )

    log_audit(
        db,
        actor_principal_id=principal.id,
        action='COUNT_SESSION_EXPORTED_CSV',
        session_id=session_id,
        ip=get_client_ip(request),
        metadata={'rows': len(variance_rows)},
    )
    db.commit()

    sio.seek(0)
    return StreamingResponse(
        iter([sio.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': f'attachment; filename=session-{session_id}-variance.csv'},
    )
